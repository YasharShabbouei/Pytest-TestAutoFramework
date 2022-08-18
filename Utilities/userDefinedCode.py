import openpyxl
import csv

def read_data_from_excel(file_address, sheetName):
    wb = openpyxl.load_workbook(filename=file_address)
    sh = wb[sheetName]
    row_max = sh.max_row
    col_max = sh.max_column
    datalist =[]

    for i in range (1, row_max+1):
        row = []
        for j in range (1, col_max+1):
            row.append(sh.cell(row=i, column=j).value)
        datalist.append(row)
    return datalist

def read_data_from_CSV(file_address):
    dataList = []
    # open csv file
    csvdata = open(file_address, "r")

    #create a reader 
    reader = csv.reader(csvdata)

    # skip the header
    next(reader)

    #add row to the list
    for row in reader:
        dataList.append(row)
    
    return dataList
